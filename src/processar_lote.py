"""
processar_lote.py

Percorre todos os PDFs de uma pasta, extrai os dados de cada um (via extrator.py)
e gera uma planilha Excel consolidada.

Uso:
    python processar_lote.py [pasta_entrada] [arquivo_saida.xlsx]

Padrão (se nenhum argumento for passado):
    pasta_entrada = pdfs_entrada/
    arquivo_saida = saida/dados_extraidos.xlsx
"""

import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from extrator import COLUNAS_SAIDA, extrair_dados_pdf


def processar_pasta(pasta_entrada: str, arquivo_saida: str):
    pasta = Path(pasta_entrada)
    pdfs = sorted(pasta.glob("*.pdf"))

    if not pdfs:
        print(f"Nenhum PDF encontrado em '{pasta_entrada}'.")
        return

    linhas = []
    erros = []

    for pdf in pdfs:
        try:
            dados = extrair_dados_pdf(str(pdf))
            linhas.append(dados)
            faltando = [c for c in COLUNAS_SAIDA if not dados.get(c)]
            if faltando:
                print(f"⚠ {pdf.name}: campos não encontrados -> {', '.join(faltando)}")
            else:
                print(f"✔ {pdf.name}: ok")
        except Exception as e:
            erros.append((pdf.name, str(e)))
            print(f"✘ {pdf.name}: ERRO -> {e}")

    _gerar_planilha(linhas, arquivo_saida)

    print(f"\n{len(linhas)} PDF(s) processado(s), {len(erros)} erro(s).")
    print(f"Planilha salva em: {arquivo_saida}")


def _gerar_planilha(linhas: list[dict], arquivo_saida: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados Extraídos"

    fonte_padrao = "Arial"

    # Cabeçalho
    header_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    header_font = Font(name=fonte_padrao, bold=True, color="FFFFFF")
    for col_idx, coluna in enumerate(COLUNAS_SAIDA, start=1):
        cel = ws.cell(row=1, column=col_idx, value=coluna)
        cel.font = header_font
        cel.fill = header_fill
        cel.alignment = Alignment(horizontal="center", vertical="center")

    # Linhas de dados
    for row_idx, dados in enumerate(linhas, start=2):
        for col_idx, coluna in enumerate(COLUNAS_SAIDA, start=1):
            valor = dados.get(coluna)
            cel = ws.cell(row=row_idx, column=col_idx, value=valor)
            cel.font = Font(name=fonte_padrao)
            if coluna == "VALOR RECIBO" and valor is not None:
                cel.value = float(valor)
                cel.number_format = '"R$" #,##0.00'
            elif coluna in ("DATA EVENTO", "DATA PROTOCOLO") and valor:
                cel.alignment = Alignment(horizontal="center")

    # Largura das colunas
    larguras = {
        "ARQUIVO": 45,
        "SEGURADO": 30,
        "Nº PROCESSO": 20,
        "VALOR RECIBO": 15,
        "NÚMERO DO CARTÃO": 22,
        "PROCEDIMENTO": 30,
        "DATA EVENTO": 14,
        "DATA PROTOCOLO": 14,
        "CPF": 16,
    }
    for col_idx, coluna in enumerate(COLUNAS_SAIDA, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = larguras.get(coluna, 18)

    ws.freeze_panes = "A2"

    Path(arquivo_saida).parent.mkdir(parents=True, exist_ok=True)
    wb.save(arquivo_saida)


if __name__ == "__main__":
    pasta_entrada = sys.argv[1] if len(sys.argv) > 1 else "pdfs_entrada"
    if len(sys.argv) > 2:
        arquivo_saida = sys.argv[2]
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo_saida = f"saida/dados_extraidos_{timestamp}.xlsx"

    processar_pasta(pasta_entrada, arquivo_saida)
